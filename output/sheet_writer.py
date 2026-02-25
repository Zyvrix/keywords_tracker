import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Column order for the output file
COLUMNS = [
    "Keyword",
    "AI Overview Present",
    "Pristyn Care in AI Overview",
    "Pristyn Care in ChatGPT",
    "Pristyn Care in Claude",
]


def init_output_file(filepath):
    """Create the output Excel file with headers if it doesn't exist."""
    if os.path.isfile(filepath):
        return  # Already exists, don't overwrite

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df = pd.DataFrame(columns=COLUMNS)
    df.to_excel(filepath, index=False, engine="openpyxl")

    # Style the header row
    wb = load_workbook(filepath)
    ws = wb.active

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    col_widths = [35, 22, 28, 25, 25]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    wb.save(filepath)
    print(f"📄 Created output file: {filepath}")


def write_result_row(filepath, row_data):
    """Append a single result row to the output Excel file."""
    wb = load_workbook(filepath)
    ws = wb.active

    next_row = ws.max_row + 1

    # Color coding
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        value = row_data.get(col_name, "")
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)

        # Apply color coding for Yes/No/N/A
        if col_idx > 1:  # Skip keyword column
            if value == "Yes":
                cell.fill = yes_fill
            elif value == "No":
                cell.fill = no_fill
            elif value == "N/A":
                cell.fill = na_fill

    wb.save(filepath)


def load_completed_keywords(filepath):
    """Load already-completed keywords from the output file (for resume support)."""
    completed = set()
    if not os.path.isfile(filepath):
        return completed

    try:
        df = pd.read_excel(filepath, engine="openpyxl")
        if "Keyword" in df.columns:
            completed = set(df["Keyword"].dropna().str.lower().str.strip())
    except Exception:
        pass

    return completed
